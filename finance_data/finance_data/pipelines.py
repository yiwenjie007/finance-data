# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://doc.scrapy.org/en/latest/topics/item-pipeline.html
from openpyxl import Workbook

class FinanceDataPipeline(object):
    def process_item(self, item, spider):
        return item

class HaiTouPipeline(object):
    def __init__(self):
        self.excel_file = 'zhaopin.xlsx'
        self.item_list = []
        self.header=['公司名称', '发布时间', '涉及城市', '文章标签', '信息地址']

    def open_spider(self, spider):
        self.wb = Workbook()
        self.sheet = self.wb.create_sheet('招聘信息', 0)
        for i in range(1, len(self.header) + 1):
            self.sheet.cell(row=1, column=i, value=self.header[i-1])

    def close_spider(self, spider):
        for i in range(2, len(self.item_list) + 2):
            for j in range(1, len(self.header) + 1):
                self.sheet.cell(row=i, column=j, value=self.get_value(dict(self.item_list[i-2]), j))
        self.wb.save(self.excel_file)
        self.wb.close()

    def process_item(self, item, spider):
        self.item_list.append(item)
        return item

    def get_value(self, item, j):
        if j == 1:
            return None if not 'company' in item.keys() else item['company']
        if j == 2:
            return None if not 'start_time' in item.keys() else item['start_time']
        if j == 3:
            return None if not 'city' in item.keys() else item['city']
        if j == 4:
            return None if not 'tag' in item.keys() else item['tag']
        if j == 5:
            return None if not 'url' in item.keys() else item['url']